"""
extract_data.py — Parse MHLW wage structure Excel files into clean CSV panels.
Creates two panels:
  1. panel_firmsize.csv: employment type × firm size × sex × year
  2. panel_industry.csv: employment type × industry × sex × year
"""

import os
import csv
import openpyxl
import xlrd

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")

# Table name mapping across years
FIRMSIZE_TABLE = {
    2014: "第７表", 2015: "第7表", 2016: "第7表", 2017: "第7表",
    2018: "第6-5表",
}
FIRMSIZE_VARIANTS_NEW = ["第6-2表", "第６－２表"]

INDUSTRY_TABLE = {
    2014: "第８表", 2015: "第8表", 2016: "第8表", 2017: "第8表",
    2018: "第6-6表",
}
INDUSTRY_VARIANTS_NEW = ["第6-3表", "第６－３表"]


def find_sheet(sheets, candidates):
    for c in candidates:
        if c in sheets:
            return c
    hw = str.maketrans('０１２３４５６７８９－', '0123456789-')
    for s in sheets:
        for c in candidates:
            if s.translate(hw) == c.translate(hw):
                return s
    return None


def read_sheet(filepath, sheetname):
    if filepath.endswith('.xls'):
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_name(sheetname)
        rows = [[ws.cell_value(i, j) for j in range(ws.ncols)] for i in range(ws.nrows)]
        wb.release_resources()
    else:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[sheetname]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
    return rows


def get_sheets(filepath):
    if filepath.endswith('.xls'):
        wb = xlrd.open_workbook(filepath)
        names = wb.sheet_names()
        wb.release_resources()
    else:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        names = wb.sheetnames
        wb.close()
    return names


def to_float(v):
    if v is None:
        return None
    s = str(v).strip().replace('＊', '').replace('…', '').replace('△', '-')
    s = s.replace('▲', '-').replace(',', '').replace('　', '').replace(' ', '')
    if s in ('', '-', '－'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def extract_sex_sections(row):
    """
    Each data row has 3 sex sections (total, male, female).
    Each section has 6 values: reg_wage, yoy, nonreg_wage, yoy, gap, prev_gap.
    Total: 18 numeric values per row.

    We extract by collecting ALL numeric values and splitting into 3 groups of 6.
    """
    nums = []
    for c in row[2:]:  # skip label columns
        v = to_float(c)
        if v is not None:
            nums.append(v)

    if len(nums) < 12:
        # Might be missing female section, try with 2 sections
        sections = []
        for i in range(0, min(len(nums), 12), 6):
            chunk = nums[i:i+6]
            if len(chunk) >= 4:
                sections.append({
                    'regular_wage': chunk[0] if chunk[0] > 100 else None,
                    'nonregular_wage': chunk[2] if len(chunk) > 2 and chunk[2] > 100 else None,
                    'wage_gap': chunk[4] if len(chunk) > 4 and 30 < chunk[4] < 100 else None,
                })
        return sections

    sections = []
    for i in range(0, min(len(nums), 18), 6):
        chunk = nums[i:i+6]
        if len(chunk) >= 5:
            reg = chunk[0] if chunk[0] > 100 else None
            nonreg = chunk[2] if chunk[2] > 100 else None
            gap = chunk[4] if 20 < chunk[4] < 100 else None
            sections.append({
                'regular_wage': reg,
                'nonregular_wage': nonreg,
                'wage_gap': gap,
            })
    return sections


# ===== FIRM SIZE PANEL =====
def parse_firmsize(year, rows):
    results = []
    firmsize_labels = {
        '大企業': 'large',
        '中企業': 'medium',
        '小企業': 'small',
    }
    sex_labels = ['total', 'male', 'female']

    for row in rows:
        label = str(row[1]).strip() if len(row) > 1 else ''
        matched = None
        for jp, en in firmsize_labels.items():
            if jp in label:
                matched = en
                break
        if matched is None:
            continue

        sections = extract_sex_sections(row)
        for j, sec in enumerate(sections):
            if j >= len(sex_labels):
                break
            results.append({
                'year': year,
                'firm_size': matched,
                'sex': sex_labels[j],
                'regular_wage': sec['regular_wage'],
                'nonregular_wage': sec['nonregular_wage'],
                'wage_gap': sec['wage_gap'],
            })

    return results


# ===== INDUSTRY PANEL =====
def parse_industry(year, rows):
    results = []
    industry_keywords = [
        ('鉱業', 'mining'),
        ('建設業', 'construction'),
        ('製造業', 'manufacturing'),
        ('電気', 'utilities'),
        ('情報通信業', 'ict'),
        ('運輸業', 'transport'),
        ('卸売業', 'wholesale_retail'),
        ('金融業', 'finance'),
        ('不動産業', 'real_estate'),
        ('学術研究', 'professional'),
        ('宿泊業', 'accommodation_food'),
        ('生活関連', 'personal_services'),
        ('教育', 'education'),
        ('医療', 'health_welfare'),
        ('複合サービス', 'compound_services'),
        ('サービス業', 'other_services'),
    ]
    sex_labels = ['total', 'male', 'female']

    for row in rows:
        label = str(row[1]).strip() if len(row) > 1 else ''
        if not label:
            continue

        matched = None
        for jp, en in industry_keywords:
            if jp in label:
                # Avoid matching 'サービス業' when it should be the catch-all
                if en == 'other_services' and any(kw in label for kw, _ in industry_keywords[:-1]):
                    continue
                matched = en
                break
        if matched is None:
            continue

        sections = extract_sex_sections(row)
        for j, sec in enumerate(sections):
            if j >= len(sex_labels):
                break
            results.append({
                'year': year,
                'industry': matched,
                'sex': sex_labels[j],
                'regular_wage': sec['regular_wage'],
                'nonregular_wage': sec['nonregular_wage'],
                'wage_gap': sec['wage_gap'],
            })

    return results


def main():
    firmsize_panel = []
    industry_panel = []

    for year in range(2014, 2025):
        ext = 'xls' if year <= 2018 else 'xlsx'
        filepath = os.path.join(DATA_DIR, f"mhlw_wage_{year}.{ext}")
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"FATAL: Missing {filepath}")

        sheets = get_sheets(filepath)

        # Firm-size table
        if year <= 2018:
            candidates = [FIRMSIZE_TABLE[year]]
        else:
            candidates = FIRMSIZE_VARIANTS_NEW
        fs_name = find_sheet(sheets, candidates)
        if fs_name:
            rows = read_sheet(filepath, fs_name)
            data = parse_firmsize(year, rows)
            firmsize_panel.extend(data)
            n_sex = len(set(r['sex'] for r in data))
            print(f"  {year} firm-size: {len(data)} rows ({n_sex} sex groups) from '{fs_name}'")

        # Industry table
        if year <= 2018:
            candidates = [INDUSTRY_TABLE[year]]
        else:
            candidates = INDUSTRY_VARIANTS_NEW
        ind_name = find_sheet(sheets, candidates)
        if ind_name:
            rows = read_sheet(filepath, ind_name)
            data = parse_industry(year, rows)
            industry_panel.extend(data)
            n_ind = len(set(r['industry'] for r in data))
            print(f"  {year} industry: {len(data)} rows ({n_ind} industries) from '{ind_name}'")

    # Write CSVs
    fs_path = os.path.join(DATA_DIR, "panel_firmsize.csv")
    with open(fs_path, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['year','firm_size','sex','regular_wage','nonregular_wage','wage_gap'])
        w.writeheader()
        w.writerows(firmsize_panel)

    ind_path = os.path.join(DATA_DIR, "panel_industry.csv")
    with open(ind_path, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['year','industry','sex','regular_wage','nonregular_wage','wage_gap'])
        w.writeheader()
        w.writerows(industry_panel)

    # Summary
    print(f"\n=== Panel Summary ===")
    print(f"Firm-size: {len(firmsize_panel)} obs")
    for sex in ['total', 'male', 'female']:
        n = sum(1 for r in firmsize_panel if r['sex'] == sex)
        print(f"  {sex}: {n} obs")

    print(f"\nIndustry: {len(industry_panel)} obs")
    for sex in ['total', 'male', 'female']:
        n = sum(1 for r in industry_panel if r['sex'] == sex)
        print(f"  {sex}: {n} obs")

    # Data validation
    print(f"\n=== Validation ===")
    for year in range(2014, 2025):
        fs_y = [r for r in firmsize_panel if r['year'] == year]
        missing_wage = sum(1 for r in fs_y if r['regular_wage'] is None)
        if missing_wage > 0:
            print(f"  {year} firm-size: {missing_wage}/{len(fs_y)} missing regular_wage")

    # Sample data check
    print(f"\n=== Sample: 2022 firm-size total ===")
    for r in firmsize_panel:
        if r['year'] == 2022 and r['sex'] == 'total':
            print(f"  {r['firm_size']}: reg={r['regular_wage']}, nonreg={r['nonregular_wage']}, gap={r['wage_gap']}")


if __name__ == '__main__':
    main()
