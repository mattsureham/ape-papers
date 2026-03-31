"""01b_parse_extra.py — Parse 2015-18 full withdrawals (Table 2) and advice data (Tables 11-14)."""

import openpyxl
import csv
import os

data_dir = "../data"
xlsx_path = os.path.join(data_dir, "rimd_2023-24.xlsx")
wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
rows_1518 = list(wb["2015-18 Data Tables"].iter_rows(values_only=True))

pot_sizes = ["<10K", "10-29K", "30-49K", "50-99K", "100-249K", "250K+"]
periods = ["H2_2015", "H1_2016", "H2_2016", "H1_2017", "H2_2017"]

# ============================================================
# Table 2: Full cash withdrawals by pot size (2015-18)
# ============================================================
# Row 21: headers show methods start at specific columns
# Annuities: col 2-11, Drawdown: col 12-21, UFPLS: col 22-31, Full cash withdrawal: col 32-41
# Each method has 5 periods × 2 cols (number, %)
# Data rows: 24-29 (6 pot sizes)

print("Parsing Table 2 (2015-18 full withdrawals by pot size)...")
records = []
for ps_idx in range(6):
    row = rows_1518[24 + ps_idx]
    # Full cash withdrawal numbers: columns 32, 34, 36, 38, 40 (every other starting at 32)
    for per_idx in range(5):
        col = 32 + per_idx * 2  # number columns
        val = row[col] if col < len(row) else None
        if val and isinstance(val, (int, float)):
            records.append({
                "period": periods[per_idx],
                "pot_size": pot_sizes[ps_idx],
                "method": "full_withdrawal",
                "count": int(val)
            })
    # Also get UFPLS: columns 22, 24, 26, 28, 30
    for per_idx in range(5):
        col = 22 + per_idx * 2
        val = row[col] if col < len(row) else None
        if val and isinstance(val, (int, float)):
            records.append({
                "period": periods[per_idx],
                "pot_size": pot_sizes[ps_idx],
                "method": "ufpls",
                "count": int(val)
            })

print(f"  Full withdrawal + UFPLS records: {len(records)}")

# Write supplement to panel
with open(os.path.join(data_dir, "panel_potsize_method_1518_supplement.csv"), "w", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=["period", "pot_size", "method", "count"])
    writer.writeheader()
    writer.writerows(records)

# ============================================================
# Tables 11-14: Advice usage by pot size
# ============================================================
# Table 14 (row 198): Full withdrawal advice
# Structure: each period has 2 cols: total at providers who specified, number where advice used

print("\nParsing advice tables...")
advice_records = []

# Table 14: Full withdrawal advice (rows 201-206)
for ps_idx in range(6):
    row = rows_1518[201 + ps_idx]
    for per_idx in range(5):
        total_col = 2 + per_idx * 2
        advised_col = 3 + per_idx * 2
        total_val = row[total_col] if total_col < len(row) else None
        advised_val = row[advised_col] if advised_col < len(row) else None
        if total_val and advised_val:
            advice_records.append({
                "period": periods[per_idx],
                "pot_size": pot_sizes[ps_idx],
                "access_method": "full_withdrawal",
                "total_at_reporting_providers": float(total_val),
                "advised": float(advised_val),
                "advice_rate": float(advised_val) / float(total_val) if float(total_val) > 0 else None
            })

# Table 12: Drawdown advice (row 168)
print("  Parsing drawdown advice (Table 12)...")
for ps_idx in range(6):
    row = rows_1518[170 + ps_idx]  # data starts ~2 rows after label
    for per_idx in range(5):
        total_col = 2 + per_idx * 2
        advised_col = 3 + per_idx * 2
        total_val = row[total_col] if total_col < len(row) else None
        advised_val = row[advised_col] if advised_col < len(row) else None
        if total_val and advised_val and isinstance(total_val, (int, float)) and isinstance(advised_val, (int, float)):
            advice_records.append({
                "period": periods[per_idx],
                "pot_size": pot_sizes[ps_idx],
                "access_method": "drawdown",
                "total_at_reporting_providers": float(total_val),
                "advised": float(advised_val),
                "advice_rate": float(advised_val) / float(total_val) if float(total_val) > 0 else None
            })

# Table 11: Annuity advice (row 154)
print("  Parsing annuity advice (Table 11)...")
for ps_idx in range(6):
    row = rows_1518[156 + ps_idx]
    for per_idx in range(5):
        total_col = 2 + per_idx * 2
        advised_col = 3 + per_idx * 2
        total_val = row[total_col] if total_col < len(row) else None
        advised_val = row[advised_col] if advised_col < len(row) else None
        if total_val and advised_val and isinstance(total_val, (int, float)) and isinstance(advised_val, (int, float)):
            advice_records.append({
                "period": periods[per_idx],
                "pot_size": pot_sizes[ps_idx],
                "access_method": "annuity",
                "total_at_reporting_providers": float(total_val),
                "advised": float(advised_val),
                "advice_rate": float(advised_val) / float(total_val) if float(total_val) > 0 else None
            })

print(f"  Total advice records: {len(advice_records)}")

with open(os.path.join(data_dir, "advice_by_potsize.csv"), "w", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=["period", "pot_size", "access_method",
                                            "total_at_reporting_providers", "advised", "advice_rate"])
    writer.writeheader()
    writer.writerows(advice_records)

# ============================================================
# Also check 2018-24 for advice data
# ============================================================
rows_1824 = list(wb["2018-24 Data Tables"].iter_rows(values_only=True))

# Find advice tables in 2018-24
print("\n2018-24 tables with 'advice':")
for i, row in enumerate(rows_1824):
    for cell in row:
        if cell and isinstance(cell, str) and 'advice' in cell.lower() and 'Table' in cell:
            print(f"  Row {i}: {cell[:100]}")

wb.close()
print("\nDone.")
