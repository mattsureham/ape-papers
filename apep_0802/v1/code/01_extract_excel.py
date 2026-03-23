#!/usr/bin/env python3
"""
01_extract_excel.py — Extract building consent data from Stats NZ Excel files.
Produces clean CSVs for R analysis.

Outputs:
  data/building_consents_by_region_type.csv  (region × dwelling_type × month)
  data/building_consents_by_ta.csv           (TA × month)
  data/population_by_ta.csv                  (TA × year)
"""

import os
import re
import sys
from pathlib import Path
import openpyxl

DATA_DIR = Path(__file__).parent.parent / "data"

# Month name to number mapping
MONTH_MAP = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
}

# Regions to include (exclude subtotals and NZ total)
VALID_REGIONS = {
    "Northland", "Auckland", "Waikato", "Bay of Plenty", "Gisborne",
    "Hawke's Bay", "Taranaki", "Manawatū-Whanganui", "Manawatu-Whanganui",
    "Wellington", "Tasman", "Nelson", "Marlborough", "West Coast",
    "Canterbury", "Otago", "Southland"
}

def clean_region(name):
    """Normalize region names across files."""
    name = re.sub(r'\(\d+\)', '', name).strip()
    name = name.replace("Manawatu-Whanganui", "Manawatū-Whanganui")
    return name

def clean_ta(name):
    """Normalize TA names."""
    name = re.sub(r'\(\d+\)', '', name).strip()
    return name

def extract_table3(wb, filename):
    """Extract dwelling-type data by region from Table 3."""
    # Find the right sheet
    sheet_name = None
    for sn in wb.sheetnames:
        if 'Table 3' == sn or 'table 3' in sn.lower():
            sheet_name = sn
            break
    if sheet_name is None:
        print(f"  WARNING: No Table 3 in {filename}")
        return []

    ws = wb[sheet_name]

    # Find year row (row 6 typically) and month row (row 7)
    years_row = None
    months_row = None
    for r in range(1, 12):
        vals = [c.value for c in list(ws.iter_rows(min_row=r, max_row=r))[0]]
        if any(isinstance(v, str) and v in MONTH_MAP for v in vals):
            months_row = r
            years_row = r - 1
            break

    if months_row is None:
        print(f"  WARNING: Could not find month headers in Table 3 of {filename}")
        return []

    # Build column-to-date mapping
    year_cells = [c.value for c in list(ws.iter_rows(min_row=years_row, max_row=years_row))[0]]
    month_cells = [c.value for c in list(ws.iter_rows(min_row=months_row, max_row=months_row))[0]]

    col_dates = {}
    current_year = None
    for i, (y, m) in enumerate(zip(year_cells, month_cells)):
        if isinstance(y, (int, float)):
            current_year = int(y)
        elif isinstance(y, str) and y.strip().isdigit():
            current_year = int(y.strip())
        if isinstance(m, str) and m.strip() in MONTH_MAP:
            if current_year:
                month_num = MONTH_MAP[m.strip()]
                col_dates[i] = f"{current_year}-{month_num:02d}-01"

    if not col_dates:
        print(f"  WARNING: No date columns found in Table 3 of {filename}")
        return []

    # Parse data rows
    records = []
    current_region = None

    for row in ws.iter_rows(min_row=months_row + 2, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row]

        # Check for region name in column A
        if vals[0] is not None and isinstance(vals[0], str) and vals[0].strip():
            candidate = clean_region(vals[0].strip())
            if candidate in VALID_REGIONS:
                current_region = candidate
            elif candidate in ("North Island", "South Island", "New Zealand",
                             "Area outside region"):
                current_region = None  # Skip subtotals
            continue

        if current_region is None:
            continue

        # Check dwelling type in column B
        if vals[1] is not None and isinstance(vals[1], str):
            dtype = vals[1].strip()
            if dtype in ("Total", "Houses", "Multi-unit homes(1)", "Multi-unit homes"):
                dtype_clean = dtype.replace("(1)", "").strip()
                if dtype_clean == "Multi-unit homes":
                    dtype_clean = "Multi-unit"

                for col_idx, date_str in col_dates.items():
                    if col_idx < len(vals):
                        val = vals[col_idx]
                        if isinstance(val, (int, float)) and val != '':
                            records.append({
                                "region": current_region,
                                "dwelling_type": dtype_clean,
                                "date": date_str,
                                "consents": int(val)
                            })

    return records


def extract_table_ta_old(wb, filename):
    """Extract TA-level data from older (pre-2022) Excel files with transposed format.
    In these files, TAs are in column A, annual data first, then monthly data further right.
    Row 7-8 contain year and month headers for the columns."""
    sheet_name = None
    for candidate in ["Table 5", "Table 6"]:
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        return []

    ws = wb[sheet_name]

    # Find the month header row (row 8 typically)
    month_row_idx = None
    year_row_idx = None
    for r in range(5, 12):
        vals = [c.value for c in list(ws.iter_rows(min_row=r, max_row=r))[0]]
        if any(isinstance(v, str) and v.strip() in MONTH_MAP for v in vals):
            month_row_idx = r
            year_row_idx = r - 1
            break

    if month_row_idx is None:
        return []

    # Build column→date mapping from year (row 7) and month (row 8) headers
    year_cells = [c.value for c in list(ws.iter_rows(min_row=year_row_idx, max_row=year_row_idx))[0]]
    month_cells = [c.value for c in list(ws.iter_rows(min_row=month_row_idx, max_row=month_row_idx))[0]]

    col_dates = {}
    current_year = None
    for i, (y, m) in enumerate(zip(year_cells, month_cells)):
        if isinstance(y, (int, float)):
            current_year = int(y)
        elif isinstance(y, str) and y.strip().isdigit():
            current_year = int(y.strip())
        if isinstance(m, str) and m.strip() in MONTH_MAP:
            if current_year:
                month_num = MONTH_MAP[m.strip()]
                col_dates[i] = f"{current_year}-{month_num:02d}-01"

    if not col_dates:
        return []

    # Parse data rows
    records = []
    # Data starts 2 rows after month headers
    for row in ws.iter_rows(min_row=month_row_idx + 2, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row]
        if vals[0] is None or not isinstance(vals[0], str):
            continue

        ta_name = clean_ta(vals[0].strip())
        if not ta_name or ta_name.startswith(("1.", "2.", "3.", "4.", "5.", "Symbol", "Source", "..", "Number")):
            continue
        if ta_name in ("New Zealand", "Percentage change", "Percentage change from same month a previous year"):
            continue
        if "percentage" in ta_name.lower() or "change" in ta_name.lower():
            continue

        for col_idx, date_str in col_dates.items():
            if col_idx < len(vals):
                val = vals[col_idx]
                if isinstance(val, (int, float)) and str(val) != '':
                    records.append({
                        "ta": ta_name,
                        "date": date_str,
                        "consents": int(val)
                    })

    return records


def extract_table6(wb, filename):
    """Extract total building consents by TA from Table 6 (or Table 5).
    Handles both newer (2022+) and older (2020-2021) formats."""
    # Try each candidate sheet until we find one with TA + monthly data
    # Require at least 3 month names in the same row to distinguish from
    # time-series format where individual months appear one per row
    for candidate in ["Table 6", "Table 5"]:
        if candidate not in wb.sheetnames:
            continue

        ws = wb[candidate]

        months_row = None
        for r in range(5, 13):
            vals = [c.value for c in list(ws.iter_rows(min_row=r, max_row=r))[0]]
            n_months = sum(1 for v in vals if isinstance(v, str) and v.strip() in MONTH_MAP)
            if n_months >= 3:
                months_row = r
                break

        if months_row is not None:
            sheet_name = candidate
            break
    else:
        print(f"  WARNING: No TA monthly data in {filename}")
        return []

    years_row = months_row - 1

    # Build column-to-date mapping
    year_cells = [c.value for c in list(ws.iter_rows(min_row=years_row, max_row=years_row))[0]]
    month_cells = [c.value for c in list(ws.iter_rows(min_row=months_row, max_row=months_row))[0]]

    col_dates = {}
    current_year = None
    for i, (y, m) in enumerate(zip(year_cells, month_cells)):
        if isinstance(y, (int, float)):
            current_year = int(y)
        elif isinstance(y, str) and y.strip().isdigit():
            current_year = int(y.strip())
        if isinstance(m, str) and m.strip() in MONTH_MAP:
            if current_year:
                month_num = MONTH_MAP[m.strip()]
                col_dates[i] = f"{current_year}-{month_num:02d}-01"

    if not col_dates:
        print(f"  WARNING: No date columns in {sheet_name} of {filename}")
        return []

    # Find data start row (skip "Number" sub-header)
    data_start = months_row + 1
    for r in range(months_row + 1, months_row + 4):
        vals = [c.value for c in list(ws.iter_rows(min_row=r, max_row=r))[0]]
        if vals[0] and isinstance(vals[0], str) and vals[0].strip().lower() == "number":
            data_start = r + 1
            break

    # Parse TA rows
    records = []
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row]

        if vals[0] is None:
            continue
        if not isinstance(vals[0], str):
            continue

        ta_name = clean_ta(vals[0].strip())

        if ta_name in ("", "New Zealand", "North Island", "South Island",
                       "Area outside territorial authority", "Percentage change",
                       "Percentage change from same month previous year",
                       "Number"):
            continue
        if ta_name.startswith(("1.", "2.", "3.", "4.", "5.", "Symbol", "Source", "..")):
            continue
        if "percentage" in ta_name.lower() or "change" in ta_name.lower():
            continue

        for col_idx, date_str in col_dates.items():
            if col_idx < len(vals):
                val = vals[col_idx]
                if isinstance(val, (int, float)) and str(val) != '':
                    records.append({
                        "ta": ta_name,
                        "date": date_str,
                        "consents": int(val)
                    })

    return records


def extract_population(filepath):
    """Extract population by TA from Stats NZ Excel.
    Format: Table 2, TAs in column A, years in row 6 (e.g., '2023', '2024 P', '2025 P')."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheet_name = None
    for sn in wb.sheetnames:
        if "Table 2" in sn:
            sheet_name = sn
            break
    if sheet_name is None:
        print(f"  WARNING: No Table 2 in {filepath.name}")
        return []

    ws = wb[sheet_name]

    # Find year headers — look for cells containing 4-digit years (with optional ' P' suffix)
    year_cols = {}
    for r in range(4, 10):
        vals = [c.value for c in list(ws.iter_rows(min_row=r, max_row=r))[0]]
        for i, v in enumerate(vals):
            if v is None:
                continue
            v_str = str(v).strip()
            year_match = re.match(r'^(20\d{2})', v_str)
            if year_match:
                year_cols[i] = int(year_match.group(1))
        if year_cols:
            year_row = r
            break

    if not year_cols:
        print(f"  WARNING: No year headers found in {filepath.name}")
        return []

    print(f"  Years found: {year_cols}")

    # Parse data rows — TAs in column A (or column A for Auckland local boards in column B)
    records = []
    data_start = year_row + 2  # Skip the "Number"/"Percent" sub-header row

    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row]

        # TA name in column A
        ta_name = None
        if vals[0] is not None and isinstance(vals[0], str) and vals[0].strip():
            ta_name = clean_ta(vals[0].strip())
        elif vals[1] is not None and isinstance(vals[1], str) and vals[1].strip():
            # Auckland local board areas are in column B — skip these for main TAs
            continue

        if not ta_name:
            continue
        if ta_name in ("Total", "New Zealand", "North Island", "South Island",
                       "Area outside territorial authority"):
            continue
        if ta_name.startswith(("1.", "2.", "3.", "Source", "Symbol", "..", "Note")):
            continue

        for col_idx, year in year_cols.items():
            if col_idx < len(vals):
                val = vals[col_idx]
                if val is not None:
                    # Handle string values like "72900" or numeric
                    if isinstance(val, str):
                        val = val.replace(",", "").strip()
                        if val.isdigit():
                            val = int(val)
                        else:
                            continue
                    if isinstance(val, (int, float)):
                        records.append({
                            "ta": ta_name,
                            "year": year,
                            "population": int(val)
                        })

    return records


def main():
    # 1. Extract building consents from all Excel files
    excel_files = sorted(DATA_DIR.glob("building-consents-issued-*.xlsx"))
    print(f"Found {len(excel_files)} building consent Excel files")

    all_region_type = []
    all_ta = []

    for f in excel_files:
        print(f"Processing: {f.name}")
        wb = openpyxl.load_workbook(f, data_only=True)

        region_records = extract_table3(wb, f.name)
        ta_records = extract_table6(wb, f.name)

        all_region_type.extend(region_records)
        all_ta.extend(ta_records)

        print(f"  Table 3: {len(region_records)} records, Table 5/6: {len(ta_records)} records")

    # Deduplicate (overlapping months across files) — keep first occurrence
    seen_region = set()
    unique_region = []
    for r in all_region_type:
        key = (r["region"], r["dwelling_type"], r["date"])
        if key not in seen_region:
            seen_region.add(key)
            unique_region.append(r)

    seen_ta = set()
    unique_ta = []
    for r in all_ta:
        key = (r["ta"], r["date"])
        if key not in seen_ta:
            seen_ta.add(key)
            unique_ta.append(r)

    # Write CSVs
    with open(DATA_DIR / "building_consents_by_region_type.csv", "w") as f:
        f.write("region,dwelling_type,date,consents\n")
        for r in sorted(unique_region, key=lambda x: (x["region"], x["dwelling_type"], x["date"])):
            f.write(f'{r["region"]},{r["dwelling_type"]},{r["date"]},{r["consents"]}\n')

    with open(DATA_DIR / "building_consents_by_ta.csv", "w") as f:
        f.write("ta,date,consents\n")
        for r in sorted(unique_ta, key=lambda x: (x["ta"], x["date"])):
            f.write(f'"{r["ta"]}",{r["date"]},{r["consents"]}\n')

    print(f"\nRegion-type records: {len(unique_region)} (unique)")
    print(f"TA records: {len(unique_ta)} (unique)")

    # Check date coverage
    if unique_region:
        region_dates = sorted(set(r["date"] for r in unique_region))
        print(f"Region-type date range: {region_dates[0]} to {region_dates[-1]} ({len(region_dates)} months)")
    if unique_ta:
        ta_dates = sorted(set(r["date"] for r in unique_ta))
        print(f"TA date range: {ta_dates[0]} to {ta_dates[-1]} ({len(ta_dates)} months)")

    # 2. Extract population
    pop_files = sorted(DATA_DIR.glob("subnational-population-*.xlsx"))
    print(f"\nFound {len(pop_files)} population files")

    all_pop = []
    for f in pop_files:
        print(f"Processing: {f.name}")
        records = extract_population(f)
        all_pop.extend(records)
        print(f"  {len(records)} records")

    # Deduplicate
    seen_pop = set()
    unique_pop = []
    for r in all_pop:
        key = (r["ta"], r["year"])
        if key not in seen_pop:
            seen_pop.add(key)
            unique_pop.append(r)

    with open(DATA_DIR / "population_by_ta.csv", "w") as f:
        f.write("ta,year,population\n")
        for r in sorted(unique_pop, key=lambda x: (x["ta"], x["year"])):
            f.write(f'"{r["ta"]}",{r["year"]},{r["population"]}\n')

    print(f"Population records: {len(unique_pop)} (unique)")

    print("\n✓ All CSVs written to data/")


if __name__ == "__main__":
    main()
