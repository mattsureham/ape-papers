#!/usr/bin/env python3
"""
Parse the DMCS 2010-2012 Excel file into a clean CSV.
The xlsx has a complex pivoted structure with 3 year-blocks and 12 monthly blocks.
R's read_excel cannot parse this; openpyxl handles it correctly.
"""
import sys
import os
import csv

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook

data_dir = os.path.join(os.path.dirname(__file__), "..", "data")
xlsx_path = os.path.join(data_dir, "dmcs_2010_2012.xlsx")
csv_path = os.path.join(data_dir, "dmcs_2010_2012_clean.csv")

if not os.path.exists(xlsx_path):
    print(f"ERROR: {xlsx_path} not found. Run 01_fetch_data.R first.")
    sys.exit(1)

print(f"Parsing {xlsx_path}...")
wb = load_workbook(xlsx_path, data_only=True)
ws = wb.active

# Column layout: year blocks start at columns 3, 63, 123 (0-indexed: 2, 62, 122)
# Each year block has 12 months x 5 columns per month
# Crime types are in rows, comunas vary by sheet structure

# Read all rows into a list
all_rows = list(ws.iter_rows(values_only=True))

# Find header row (contains "COMUNA" or similar)
header_row_idx = None
for i, row in enumerate(all_rows):
    for cell in row:
        if cell and str(cell).strip().upper() == "COMUNA":
            header_row_idx = i
            break
    if header_row_idx is not None:
        break

if header_row_idx is None:
    print("ERROR: Could not find header row with 'COMUNA'")
    sys.exit(1)

# Crime type columns for each year block
# Structure: for each month, there are columns for each crime type
# The exact structure varies - we need to identify the pattern

# Simpler approach: read the known structure
# Year starts at specific column offsets
year_starts = {2010: 2, 2011: 62, 2012: 122}

crime_types = [
    "Homicidio", "Robo_Intimidacion", "Robo_Violencia", "Robo_Sorpresa",
    "Robo_Vehiculo_Motorizado", "Robo_Accesorio_Vehiculo",
    "Robo_Lugar_Habitado", "Robo_Lugar_No_Habitado", "Otros_Robos_Fuerza",
    "Hurto", "Lesiones", "Violacion", "VIF", "Ley_Drogas"
]

# Each crime has 4 sub-columns: CASOS, DENUNCIAS, DETENCION, DETENIDOS
# We want CASOS (first sub-column for each crime)
cols_per_crime = 4
cols_per_month = len(crime_types) * cols_per_crime

records = []
data_start = header_row_idx + 1

for row_idx in range(data_start, len(all_rows)):
    row = all_rows[row_idx]
    comuna = row[0] if row[0] else row[1]
    if not comuna or str(comuna).strip() == "":
        continue
    comuna = str(comuna).strip()
    if comuna.upper() in ("TOTAL", "TOTAL GENERAL", "TOTALES"):
        continue

    for year, start_col in year_starts.items():
        for month in range(12):
            month_offset = start_col + month * cols_per_month
            record = {"comuna": comuna, "year": year, "month": month + 1}

            for ci, crime in enumerate(crime_types):
                col_idx = month_offset + ci * cols_per_crime
                if col_idx < len(row):
                    val = row[col_idx]
                    try:
                        record[crime] = int(val) if val is not None else 0
                    except (ValueError, TypeError):
                        record[crime] = 0
                else:
                    record[crime] = 0

            # Only add if we have at least some non-zero data
            total = sum(record.get(c, 0) for c in crime_types)
            if total > 0 or any(record.get(c, 0) != 0 for c in crime_types):
                records.append(record)

print(f"Parsed {len(records)} records")

# Write CSV
fieldnames = ["comuna", "year", "month"] + crime_types
with open(csv_path, "w", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(records)

print(f"Saved to {csv_path}")
